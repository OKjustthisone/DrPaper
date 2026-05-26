import os
import uuid
import csv
import io
from pathlib import Path
from typing import Optional
from app.config import UPLOADS_DIR


def parse_file(file_path: str) -> list[dict]:
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return _parse_pdf(file_path)
    elif ext == ".docx":
        return _parse_docx(file_path)
    elif ext in (".md", ".txt", ".markdown"):
        return _parse_text(file_path)
    elif ext == ".csv":
        return _parse_csv(file_path)
    elif ext == ".xlsx":
        return _parse_xlsx(file_path)
    elif ext in (".png", ".jpg", ".jpeg", ".webp"):
        return _parse_image_ocr(file_path)
    else:
        return [{"text": f"Unsupported file type: {ext}", "metadata": {}}]


def _parse_pdf(file_path: str) -> list[dict]:
    chunks = []
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text and text.strip():
                    chunks.append({
                        "text": text.strip(),
                        "page": i + 1,
                        "metadata": {"page": i + 1, "source": Path(file_path).name}
                    })
    except ImportError:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                if text and text.strip():
                    chunks.append({
                        "text": text.strip(),
                        "page": i + 1,
                        "metadata": {"page": i + 1, "source": Path(file_path).name}
                    })
        except ImportError:
            chunks.append({"text": "PDF parsing unavailable. Install pdfplumber.", "metadata": {}})
    return chunks


def _parse_docx(file_path: str) -> list[dict]:
    try:
        from docx import Document
        doc = Document(file_path)
        full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return [{"text": full_text, "metadata": {"source": Path(file_path).name}}]
    except ImportError:
        return [{"text": "DOCX parsing unavailable. Install python-docx.", "metadata": {}}]


def _parse_text(file_path: str) -> list[dict]:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    return [{"text": text, "metadata": {"source": Path(file_path).name}}]


def _parse_csv(file_path: str) -> list[dict]:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        return [{"text": "(Empty CSV)", "metadata": {}}]
    columns = list(rows[0].keys())
    text_lines = [
        f"CSV File: {Path(file_path).name}",
        f"Columns: {', '.join(columns)}",
        f"Row count: {len(rows)}",
        "",
    ]
    for i, row in enumerate(rows):
        line = " | ".join(f"{col}: {row.get(col, '')}" for col in columns)
        text_lines.append(f"Row {i+1}: {line}")
    return [{"text": "\n".join(text_lines), "metadata": {"source": Path(file_path).name, "columns": columns}}]


def _parse_xlsx(file_path: str) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        chunks = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c) if c else "" for c in rows[0]]
            text_lines = [
                f"Sheet: {sheet_name}",
                f"Columns: {', '.join(headers)}",
                f"Row count: {len(rows) - 1}",
                "",
            ]
            for i, row in enumerate(rows[1:], 1):
                line = " | ".join(f"{h}: {v}" for h, v in zip(headers, row) if v is not None)
                text_lines.append(f"Row {i}: {line}")
            chunks.append({
                "text": "\n".join(text_lines),
                "metadata": {"source": Path(file_path).name, "sheet": sheet_name}
            })
        return chunks
    except ImportError:
        return [{"text": "XLSX parsing unavailable. Install openpyxl.", "metadata": {}}]


def _parse_image_ocr(file_path: str) -> list[dict]:
    try:
        from PIL import Image
        import pytesseract
        img = Image.open(file_path)
        text = pytesseract.image_to_string(img, lang="eng+chi_sim")
        return [{"text": text, "metadata": {"source": Path(file_path).name, "ocr": True}}]
    except ImportError:
        return [{"text": f"[Image: {Path(file_path).name}] (OCR unavailable)", "metadata": {"source": Path(file_path).name}}]
    except Exception as e:
        return [{"text": f"[Image: {Path(file_path).name}] (OCR error: {e})", "metadata": {}}]


def save_upload(file_content: bytes, filename: str) -> str:
    unique_name = f"{uuid.uuid4().hex}_{filename}"
    file_path = os.path.join(UPLOADS_DIR, unique_name)
    with open(file_path, "wb") as f:
        f.write(file_content)
    return file_path
